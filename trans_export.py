#!/usr/local/bin/python
# -*- coding:utf-8 -*-
import sys
import os
import re
from openpyxl import Workbook
from collections import defaultdict
from collections import namedtuple
import pymssql


server = 'BJ-SQL02.fihtdc.com'
user='nString'
password='Nstring123456'

class TranslateItem(object):
    def __init__(self, account, project, feature, textid, langcode,
            translation):
        self.account = account
        self.project = project
        self.feature = feature
        self.textid = textid
        self.langcode = langcode
        self.translation = translation

class WhereClause(object):
    def __init__(self):
        self.clauses = ''

    def add(self, field, values):
        if len(values)<=0: # do nothing empty values
            return

        clause = '{0} IN ({1})'.format(field, ', '.join(map(repr,values)))


        if self.clauses:
            self.clauses = ' {0} AND {1}'.format(self.clauses, clause)
        else:
            self.clauses = ' WHERE {0}'.format(clause)


Account = namedtuple('Account', ['id', 'name'])

class DBWrapper(object):
    def __init__(self):
        self.conn = None
        self.connect()

    def connect(self):
        self.conn = pymssql.connect(server, user, password, 
                'nstring',charset='utf8')
    def close(self):
        self.conn.close()

    def __enter__(self):
        return self
    def __exit__(self, type, value, traceback):
        self.close()

    def select_translation(self, where_clause):
        with self.conn.cursor(as_dict=True) as cursor:
            sql = "select Text, Account, Project, Feature, Language, Lv1 \
            from TranslationView" + where_clause

            cursor.execute(sql)

            lines = []
            translations = []
            for row in cursor:
                trans_item = \
                TranslateItem(row['Account'], row['Project'],
                row['Feature'],row['Text'], row['Language'],
                row['Lv1'])

                translations.append(trans_item)

            return translations

    def select_account(self):
        with self.conn.cursor(as_dict=True) as cursor:
            sql = "select id, name from Customer"
            cursor.execute(sql)

            accounts = []
            for row in cursor:
                accounts.append(Account(row['id'],row['name']))

            return accounts

   
    def select_language(self, account_id):
        with self.conn.cursor(as_dict=True) as cursor:
            sql = "select Language from ViewLanguageSetLanguage" + \
                    " where account_id=" + str(account_id)
            cursor.execute(sql)

            langs = []
            for row in cursor:
                langs.append(row['Language'])

            return langs

    def select_project(self, account_id):
        with self.conn.cursor(as_dict=True) as cursor:
            sql = "select name from Project where account_id =" + \
            str(account_id)
            
            cursor.execute(sql)

            projects = []
            for row in cursor:
                projects.append(row['name'])

            return projects





class TransExport(object):
    def _get_lancode_set(self, translations):
        '''
        :type:
            translation: list of TranslateItem
        :rtype: 
            set contains all langcode
        '''
        lancode_set = set()
        for trans in translations:
            lancode_set.add(trans.langcode)

        return lancode_set

    def _get_id_map(self, translations):
        '''
        :type:
            translation: list of TranslateItem
        :rtype: 
            dict which map textid and all its translation
        '''
        id_map = defaultdict(list)
        for trans in translations:
            id_map[trans.textid].append(trans)

        return id_map


    def save_translation(self, translations, xlsx):
        id_map = self._get_id_map(translations)
        lancode_set = self._get_lancode_set(translations)

        wb = Workbook()
        ws = wb.active

        # create map from langcode to column index in xlsx
        # fill title row 
        cur_row = 1
        cur_col = 1
        ws.cell(row=cur_row, column=cur_col).value = 'RefName' 
        col_map = {'RefName':cur_col}
        cur_col += 1
        ws.cell(row=cur_row, column=cur_col).value = 'ModOP' 
        col_map['ModOP'] = cur_col

        for code in sorted(lancode_set):
            cur_col += 1
            ws.cell(row=cur_row, column=cur_col).value = code
            col_map[code] = cur_col 

        # fill translation row
        for textid in sorted(id_map.keys()):
            cur_row += 1
            cur_col = col_map['RefName']
            ws.cell(row=cur_row, column=cur_col).value = textid
            cur_col = col_map['ModOP']
            ws.cell(row=cur_row, column=cur_col).value = \
                    id_map[textid][0].feature
            for trans in id_map[textid]:
                cur_col = col_map[trans.langcode]
                ws.cell(row=cur_row, column=cur_col).value = trans.translation

        wb.save(xlsx)




def get_values(fname):
    '''
    rtype: list of non-blanck lines in file fname
    '''
    values = []
    with open(fname,'r',encoding='utf8') as f:
        for line in f.readlines():
            line = line.strip()
            if line:
                values.append(line)

    return values


def export_argon():
    where_clause = WhereClause()
    where_clause.add('Account', ['KAIOS_UPDATE'])
    text_ids = get_values('data/textids.txt')
    where_clause.add('Text', text_ids)
    langs = get_values('data/langs.txt')
    where_clause.add('Language', langs)

    with DBWrapper() as db:
        translation = db.select_translation(where_clause.clauses)
        print(where_clause.clauses)
        export = TransExport()
        export.save_translation(translations, 'data/argon_translations.xlsx')

def export_dougla():
    where_clause = WhereClause()
    where_clause.add('Account', ['S30Plus'])
    where_clause.add('Project', ['douglas8cu bit'])
    langs = ('English-GB', 'Chinese-HK', 'Chinese-CN', 'Russian', 'German', 'French-CA', 'Arabic', 'Persian', 'Malay', 'Hindi', 'Finnish')
    where_clause.add('Language', langs)

    with DBWrapper() as db:
        translation = db.select_translation(where_clause.clauses)
        print(where_clause.clauses)
        export = TransExport()
        export.save_translation(translations, 'data/douglas8cu_translations.xlsx')


if __name__ == '__main__':
    export_dougla()
    # export_argon()

